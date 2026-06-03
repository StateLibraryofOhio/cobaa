import io

from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from cart.cart import Cart
from orders.utils import order_create


def books_to_excel(request):
    cart = Cart(request)
    order = order_create(request)
    if not order:
        messages.add_message(request, messages.ERROR, message='Oops... Something went wrong.')
        return HttpResponse('Oops... Something went wrong.', status=400)

    wb = Workbook()
    ws = wb.active
    ws.title = 'My Book List'

    # Header row
    headers = ['Title', 'Authors / Illustrators', 'Quantity']
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(fill_type='solid', fgColor='2563EB')
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')

    # Data rows
    row_num = 2
    total_items = 0
    for item in cart:
        book = item['book']
        authors = ', '.join(str(a) for a in book.authors.all())
        qty = item['quantity']
        total_items += qty
        ws.cell(row=row_num, column=1, value=book.title)
        ws.cell(row=row_num, column=2, value=authors)
        ws.cell(row=row_num, column=3, value=qty)
        row_num += 1

    # Summary rows
    ws.cell(row=row_num + 1, column=1, value='Total Items').font = Font(bold=True)
    ws.cell(row=row_num + 1, column=3, value=total_items).font = Font(bold=True)
    ws.cell(row=row_num + 2, column=1, value='Retrieval Code').font = Font(bold=True)
    ws.cell(row=row_num + 2, column=2, value=order.id).font = Font(bold=True)

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 50
    ws.column_dimensions[get_column_letter(2)].width = 35
    ws.column_dimensions[get_column_letter(3)].width = 12

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'{order.inst_name} COBAA Selections.xlsx'
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
